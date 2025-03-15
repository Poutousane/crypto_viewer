import streamlit as st
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# Configuration de la page
st.set_page_config(
    page_title="Finance Viewer",
    layout="wide",
    initial_sidebar_state="expanded"
)


# Fonction pour définir le mode sombre/clair
def set_theme():
    if st.session_state.theme_mode == "Sombre":
        # CSS pour le mode sombre
        st.markdown("""
        <style>
            .stApp {
                background-color: #0E1117;
                color: #FAFAFA;
            }
            .css-1d391kg, .css-12oz5g7 {
                background-color: #262730;
                color: #FAFAFA;
            }
            .stTabs [data-baseweb="tab-list"] {
                background-color: #262730;
            }
            .stTabs [data-baseweb="tab"] {
                color: #FAFAFA;
            }
            .stDataFrame {
                background-color: #262730;
            }
        </style>
        """, unsafe_allow_html=True)
    else:
        # CSS pour le mode clair (par défaut)
        st.markdown("""
        <style>
            .stApp {
                background-color: white;
                color: black;
            }
            .css-1d391kg, .css-12oz5g7 {
                background-color: white;
            }
            .stTabs [data-baseweb="tab-list"] {
                background-color: white;
            }
            .stTabs [data-baseweb="tab"] {
                color: black;
            }
        </style>
        """, unsafe_allow_html=True)


# Initialisation de l'état de session pour le thème
if 'theme_mode' not in st.session_state:
    st.session_state.theme_mode = "Clair"

# Barre latérale pour les configurations
st.sidebar.title("Paramètres")

# Bouton de thème dans la barre latérale
theme_options = ["Clair", "Sombre"]
selected_theme = st.sidebar.selectbox(
    "Thème",
    options=theme_options,
    index=theme_options.index(st.session_state.theme_mode)
)

if selected_theme != st.session_state.theme_mode:
    st.session_state.theme_mode = selected_theme
    set_theme()
    st.experimental_rerun()

# Appliquer le thème
set_theme()

# Titre de l'application
st.title("Finance Viewer")

# Définition des catégories d'actions par secteur
stock_categories = {
    "Tech": {
        "Apple": "AAPL",
        "Microsoft": "MSFT",
        "Google": "GOOGL",
        "Amazon": "AMZN",
        "Tesla": "TSLA",
        "Meta": "META",
        "NVIDIA": "NVDA",
        "Adobe": "ADBE",
        "Intel": "INTC",
        "IBM": "IBM",
        "Cisco": "CSCO",
        "Oracle": "ORCL",
        "Salesforce": "CRM",
        "AMD": "AMD",
        "PayPal": "PYPL",
    },
    "Banques & Finance": {
        "JPMorgan Chase": "JPM",
        "Bank of America": "BAC",
        "Wells Fargo": "WFC",
        "Goldman Sachs": "GS",
        "Morgan Stanley": "MS",
        "Visa": "V",
        "Mastercard": "MA",
        "American Express": "AXP",
    },
    "Consommation": {
        "Walmart": "WMT",
        "Coca-Cola": "KO",
        "PepsiCo": "PEP",
        "McDonald's": "MCD",
        "Nike": "NKE",
        "Disney": "DIS",
        "Home Depot": "HD",
        "Starbucks": "SBUX",
        "Procter & Gamble": "PG",
        "Netflix": "NFLX",
    },
    "Santé & Pharmacie": {
        "Johnson & Johnson": "JNJ",
        "Pfizer": "PFE",
        "Merck": "MRK",
        "UnitedHealth": "UNH",
        "Abbott Labs": "ABT",
        "Eli Lilly": "LLY",
        "Amgen": "AMGN",
        "Bristol-Myers Squibb": "BMY",
    },
    "Énergie": {
        "Exxon Mobil": "XOM",
        "Chevron": "CVX",
        "ConocoPhillips": "COP",
        "Shell": "SHEL",
        "BP": "BP",
    },
    "Automobile": {
        "Ford": "F",
        "General Motors": "GM",
        "Toyota": "TM",
        "Honda": "HMC",
        "Volkswagen": "VWAGY",
    },
    "Télécommunications": {
        "AT&T": "T",
        "Verizon": "VZ",
        "T-Mobile": "TMUS",
        "Comcast": "CMCSA"
    }
}

# Conversion à un dictionnaire plat pour faciliter la recherche
stock_assets = {}
for category, assets in stock_categories.items():
    for name, ticker in assets.items():
        stock_assets[name] = ticker

# Listes des autres types d'actifs
crypto_assets = {
    "Bitcoin": "BTC-USD",
    "Ethereum": "ETH-USD",
    "Binance Coin": "BNB-USD",
    "Solana": "SOL-USD",
    "XRP": "XRP-USD",
    "Cardano": "ADA-USD",
    "Dogecoin": "DOGE-USD",
    "Polkadot": "DOT-USD"
}

currency_assets = {
    "EUR/USD": "EURUSD=X",
    "GBP/USD": "GBPUSD=X",
    "USD/JPY": "USDJPY=X",
    "USD/CAD": "USDCAD=X",
    "AUD/USD": "AUDUSD=X",
    "USD/CHF": "USDCHF=X",
    "NZD/USD": "NZDUSD=X",
    "EUR/GBP": "EURGBP=X"
}

resource_assets = {
    "Or": "GC=F",
    "Argent": "SI=F",
    "Pétrole brut": "CL=F",
    "Gaz naturel": "NG=F",
    "Cuivre": "HG=F",
    "Blé": "ZW=F",
    "Maïs": "ZC=F"
}

index_assets = {
    "S&P 500": "^GSPC",
    "NASDAQ Composite": "^IXIC",
    "Dow Jones": "^DJI",
    "Russell 2000": "^RUT",
    "CAC 40": "^FCHI",
    "DAX": "^GDAXI",
    "FTSE 100": "^FTSE",
    "Nikkei 225": "^N225",
    "Hang Seng": "^HSI",
    "ASX 200": "^AXJO",
    "IBEX 35": "^IBEX",
    "FTSE MIB": "FTSEMIB.MI",
    "KOSPI": "^KS11",
    "TSX Composite": "^GSPTSE"
}


# Fonction pour obtenir les données fondamentales
def get_fundamentals(ticker):
    try:
        stock = yf.Ticker(ticker)

        # Récupérer les informations de base
        info = stock.info

        # Préparer les données fondamentales
        fundamentals = {}

        # Ajouter les données clés si elles existent
        keys_to_check = [
            'marketCap', 'trailingPE', 'forwardPE', 'dividendYield',
            'beta', 'bookValue', 'priceToBook', 'returnOnEquity',
            'trailingEps', 'forwardEps', 'pegRatio', 'enterpriseToEbitda'
        ]

        for key in keys_to_check:
            if key in info:
                fundamentals[key] = info[key]

        return fundamentals

    except Exception as e:
        return {"error": str(e)}


# Fonction pour formater les grandes valeurs numériques
def format_large_number(num):
    if pd.isna(num) or num is None:
        return "N/A"

    if isinstance(num, str):
        return num

    if num >= 1e12:
        return f"{num / 1e12:.2f}T"
    elif num >= 1e9:
        return f"{num / 1e9:.2f}B"
    elif num >= 1e6:
        return f"{num / 1e6:.2f}M"
    elif num >= 1e3:
        return f"{num / 1e3:.2f}K"
    else:
        return f"{num:.2f}"


# Fonction pour créer un Excel avec la colonne variation formatée en pourcentage
def create_excel(data, sheet_name="Data"):
    output = io.BytesIO()

    # Formater les données pour l'export
    export_data = data.copy()

    # Calculer la variation quotidienne
    export_data['Variation (%)'] = export_data['Close'].pct_change() * 100

    # Convertir les indices en dates au format YYYY-MM-DD
    export_data.index = [d.strftime('%Y-%m-%d') for d in export_data.index]
    export_data.index.name = 'Date'

    # Réorganiser les colonnes
    export_data = export_data[['Close', 'High', 'Low', 'Open', 'Variation (%)', 'Volume']]

    # Renommer les colonnes
    export_data.columns = ['Price', 'High', 'Low', 'Open', 'Variation (%)', 'Volume']

    # Créer et configurer manuellement le fichier Excel
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name

    # Ajouter la ligne d'en-tête
    headers = ['Date'] + list(export_data.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Ajouter les données
    for row_idx, (date_idx, row_data) in enumerate(export_data.iterrows(), start=2):
        # Ajouter la date
        ws.cell(row=row_idx, column=1, value=date_idx)

        # Ajouter les autres colonnes
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Pour la colonne Variation (%), appliquer le format pourcentage
            if col_idx == 6:  # 6 est l'index pour Variation (%)
                if pd.notna(value):
                    cell.value = value / 100  # Convertir en décimal pour Excel
                    cell.number_format = '0.00%'  # Format pourcentage avec 2 décimales

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    workbook.save(output)
    processed_data = output.getvalue()
    return processed_data


# Création des onglets principaux pour types d'actifs
tab1, tab2, tab3, tab4, tab5 = st.tabs(["Crypto", "Actions", "Devises", "Ressources", "Indices"])


# Fonction pour afficher les crypto, devises, ressources et indices
def display_standard_asset_data(assets, tab_key):
    # Barre de recherche
    search_term = st.text_input("Recherche rapide", key=f"search_{tab_key}")

    # Filtrer les actifs en fonction du terme de recherche
    filtered_assets = {k: v for k, v in assets.items() if search_term.lower() in k.lower()}

    if not filtered_assets:
        st.warning(f"Aucun actif ne correspond à '{search_term}'")
        return

    col1, col2, col3 = st.columns(3)

    with col1:
        selected_asset = st.selectbox("Choisissez un actif", list(filtered_assets.keys()), key=f"select_{tab_key}")

    with col2:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=365)
        start_date_input = st.date_input("Date de début", value=start_date, key=f"start_{tab_key}")

    with col3:
        end_date_input = st.date_input("Date de fin", value=end_date, key=f"end_{tab_key}")

    # Récupération des données
    ticker_symbol = filtered_assets[selected_asset]
    try:
        # Récupérer les données avec un intervalle quotidien explicite
        data = yf.download(ticker_symbol, start=start_date_input, end=end_date_input, interval="1d")

        if data.empty:
            st.error(f"Aucune donnée disponible pour {selected_asset} dans la période sélectionnée.")
        else:
            # S'assurer que les données ne sont pas vides
            if not data.empty and len(data) > 0:
                # Extraction des valeurs
                latest_close_value = float(data['Close'].iloc[-1])
                first_close_value = float(data['Close'].iloc[0])
                variation_value = ((latest_close_value - first_close_value) / first_close_value) * 100
                latest_volume_value = float(data['Volume'].iloc[-1])

                # Affichage des indicateurs clés
                st.subheader("Indicateurs clés")

                metrics_col1, metrics_col2, metrics_col3 = st.columns(3)

                with metrics_col1:
                    formatted_close = f"${latest_close_value:.2f}"
                    st.metric("Prix de clôture", formatted_close)

                with metrics_col2:
                    formatted_variation = f"{variation_value:.2f}%"
                    formatted_delta = f"{variation_value:.2f}%"
                    st.metric("Variation", formatted_variation, delta=formatted_delta)

                with metrics_col3:
                    if latest_volume_value >= 1e9:
                        vol_str = f"{latest_volume_value / 1e9:.2f} G"
                    elif latest_volume_value >= 1e6:
                        vol_str = f"{latest_volume_value / 1e6:.2f} M"
                    elif latest_volume_value >= 1e3:
                        vol_str = f"{latest_volume_value / 1e3:.2f} k"
                    else:
                        vol_str = f"{latest_volume_value:.2f}"
                    st.metric("Volume (dernier jour)", vol_str)

                # Tableau des données
                st.subheader("Données historiques")

                # Calculer la variation quotidienne
                data['Daily_Change'] = data['Close'].pct_change() * 100

                # Créer un DataFrame pour l'affichage avec l'index recréé sans l'heure
                display_data = pd.DataFrame(index=[d.date() for d in data.index])

                # Convertir les valeurs numpy en valeurs Python natives
                open_values = [float(x) for x in data['Open'].to_numpy()]
                high_values = [float(x) for x in data['High'].to_numpy()]
                low_values = [float(x) for x in data['Low'].to_numpy()]
                close_values = [float(x) for x in data['Close'].to_numpy()]

                # Formater les prix
                display_data['Open'] = [f"${x:.2f}" for x in open_values]
                display_data['High'] = [f"${x:.2f}" for x in high_values]
                display_data['Low'] = [f"${x:.2f}" for x in low_values]
                display_data['Close'] = [f"${x:.2f}" for x in close_values]

                # Formater la variation avec gestion des NaN
                daily_changes = []
                for x in data['Daily_Change'].to_numpy():
                    if pd.isna(x):
                        daily_changes.append("N/A")
                    else:
                        daily_changes.append(f"{float(x):.2f}%")
                display_data['Variation (%)'] = daily_changes

                # Formater le volume
                volumes = []
                for vol in data['Volume'].to_numpy():
                    vol = float(vol)
                    if vol >= 1e9:
                        volumes.append(f"{vol / 1e9:.2f} G")
                    elif vol >= 1e6:
                        volumes.append(f"{vol / 1e6:.2f} M")
                    elif vol >= 1e3:
                        volumes.append(f"{vol / 1e3:.2f} k")
                    else:
                        volumes.append(f"{vol:.2f}")
                display_data['Volume'] = volumes

                # Affichage du tableau avec filtres
                st.dataframe(display_data)

                # Créer un excel avec les colonnes inversées et le format pourcentage pour la variation
                excel_data = create_excel(data, selected_asset)

                st.download_button(
                    label="Télécharger les données (XLSX)",
                    data=excel_data,
                    file_name=f'{selected_asset}_{start_date_input}_{end_date_input}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key=f"download_{tab_key}"
                )
            else:
                st.error(f"Aucune donnée n'a été récupérée pour {selected_asset}.")

    except Exception as e:
        st.error(f"Une erreur s'est produite lors de la récupération des données : {e}")
        import traceback
        st.error(f"Traceback détaillé: {traceback.format_exc()}")


# Fonction spéciale pour afficher les actions (avec le filtrage par secteur et les données fondamentales)
def display_stock_data():
    # Barre de recherche
    search_term = st.text_input("Recherche rapide", key="search_stock")

    # Option de filtrage par secteur
    all_sectors = list(stock_categories.keys())
    all_sectors.insert(0, "Tous les secteurs")

    selected_sector = st.selectbox(
        "Filtrer par secteur",
        options=all_sectors,
        key="sector_filter"
    )

    # Filtrer les actions en fonction du secteur et du terme de recherche
    filtered_stocks = {}

    if selected_sector == "Tous les secteurs":
        # Filtrer toutes les actions basées sur la recherche
        filtered_stocks = {k: v for k, v in stock_assets.items() if search_term.lower() in k.lower()}
    else:
        # Filtrer les actions du secteur sélectionné
        sector_stocks = stock_categories[selected_sector]
        filtered_stocks = {k: v for k, v in sector_stocks.items() if search_term.lower() in k.lower()}

    if not filtered_stocks:
        if search_term:
            st.warning(f"Aucune action ne correspond à '{search_term}' dans le secteur sélectionné.")
        else:
            st.warning(f"Aucune action disponible dans le secteur sélectionné.")
        return

    # Sélection d'une action
    col1, col2, col3 = st.columns(3)

    with col1:
        selected_asset = st.selectbox("Choisissez une action", list(filtered_stocks.keys()), key="select_stock")

    with col2:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=365)
        start_date_input = st.date_input("Date de début", value=start_date, key="start_stock")

    with col3:
        end_date_input = st.date_input("Date de fin", value=end_date, key="end_stock")

    # Récupération des données
    ticker_symbol = filtered_stocks[selected_asset]
    try:
        # Récupérer les données avec un intervalle quotidien explicite
        data = yf.download(ticker_symbol, start=start_date_input, end=end_date_input, interval="1d")

        if data.empty:
            st.error(f"Aucune donnée disponible pour {selected_asset} dans la période sélectionnée.")
        else:
            # S'assurer que les données ne sont pas vides
            if not data.empty and len(data) > 0:
                # Extraction des valeurs
                latest_close_value = float(data['Close'].iloc[-1])
                first_close_value = float(data['Close'].iloc[0])
                variation_value = ((latest_close_value - first_close_value) / first_close_value) * 100
                latest_volume_value = float(data['Volume'].iloc[-1])

                # Récupérer les données fondamentales
                fundamentals = get_fundamentals(ticker_symbol)

                # Structure en deux colonnes pour les données fondamentales et le graphique
                col_fund, col_metrics = st.columns([1, 2])

                with col_fund:
                    st.subheader("Données fondamentales")

                    if "error" in fundamentals:
                        st.warning("Impossible de récupérer les données fondamentales")
                    else:
                        fund_data = []

                        # Formatter les données fondamentales
                        if "marketCap" in fundamentals:
                            fund_data.append(
                                ("Capitalisation boursière", format_large_number(fundamentals.get("marketCap"))))

                        if "trailingPE" in fundamentals:
                            fund_data.append(("Ratio P/E (trailing)", f"{fundamentals.get('trailingPE', 'N/A'):.2f}"))

                        if "forwardPE" in fundamentals:
                            fund_data.append(("Ratio P/E (forward)", f"{fundamentals.get('forwardPE', 'N/A'):.2f}"))

                        if "dividendYield" in fundamentals and fundamentals["dividendYield"] is not None:
                            div_yield = fundamentals["dividendYield"] * 100
                            fund_data.append(("Rendement du dividende", f"{div_yield:.2f}%"))

                        if "beta" in fundamentals:
                            fund_data.append(("Bêta", f"{fundamentals.get('beta', 'N/A'):.2f}"))

                        if "priceToBook" in fundamentals:
                            fund_data.append(
                                ("Ratio prix/valeur comptable", f"{fundamentals.get('priceToBook', 'N/A'):.2f}"))

                        if "returnOnEquity" in fundamentals and fundamentals["returnOnEquity"] is not None:
                            roe = fundamentals["returnOnEquity"] * 100
                            fund_data.append(("Rendement des capitaux propres", f"{roe:.2f}%"))

                        # Créer et afficher un dataframe pour les données fondamentales
                        if fund_data:
                            fund_df = pd.DataFrame(fund_data, columns=["Métrique", "Valeur"])
                            st.table(fund_df)
                        else:
                            st.info("Aucune donnée fondamentale disponible")

                with col_metrics:
                    # Affichage des indicateurs clés
                    st.subheader("Indicateurs clés")

                    metrics_col1, metrics_col2, metrics_col3 = st.columns(3)

                    with metrics_col1:
                        formatted_close = f"${latest_close_value:.2f}"
                        st.metric("Prix de clôture", formatted_close)

                    with metrics_col2:
                        formatted_variation = f"{variation_value:.2f}%"
                        formatted_delta = f"{variation_value:.2f}%"
                        st.metric("Variation", formatted_variation, delta=formatted_delta)

                    with metrics_col3:
                        if latest_volume_value >= 1e9:
                            vol_str = f"{latest_volume_value / 1e9:.2f} G"
                        elif latest_volume_value >= 1e6:
                            vol_str = f"{latest_volume_value / 1e6:.2f} M"
                        elif latest_volume_value >= 1e3:
                            vol_str = f"{latest_volume_value / 1e3:.2f} k"
                        else:
                            vol_str = f"{latest_volume_value:.2f}"
                        st.metric("Volume (dernier jour)", vol_str)

                # Tableau des données
                st.subheader("Données historiques")

                # Calculer la variation quotidienne
                data['Daily_Change'] = data['Close'].pct_change() * 100

                # Créer un DataFrame pour l'affichage avec l'index recréé sans l'heure
                display_data = pd.DataFrame(index=[d.date() for d in data.index])

                # Convertir les valeurs numpy en valeurs Python natives
                open_values = [float(x) for x in data['Open'].to_numpy()]
                high_values = [float(x) for x in data['High'].to_numpy()]
                low_values = [float(x) for x in data['Low'].to_numpy()]
                close_values = [float(x) for x in data['Close'].to_numpy()]

                # Formater les prix
                display_data['Open'] = [f"${x:.2f}" for x in open_values]
                display_data['High'] = [f"${x:.2f}" for x in high_values]
                display_data['Low'] = [f"${x:.2f}" for x in low_values]
                display_data['Close'] = [f"${x:.2f}" for x in close_values]

                # Formater la variation avec gestion des NaN
                daily_changes = []
                for x in data['Daily_Change'].to_numpy():
                    if pd.isna(x):
                        daily_changes.append("N/A")
                    else:
                        daily_changes.append(f"{float(x):.2f}%")
                display_data['Variation (%)'] = daily_changes

                # Formater le volume
                volumes = []
                for vol in data['Volume'].to_numpy():
                    vol = float(vol)
                    if vol >= 1e9:
                        volumes.append(f"{vol / 1e9:.2f} G")
                    elif vol >= 1e6:
                        volumes.append(f"{vol / 1e6:.2f} M")
                    elif vol >= 1e3:
                        volumes.append(f"{vol / 1e3:.2f} k")
                    else:
                        volumes.append(f"{vol:.2f}")
                display_data['Volume'] = volumes

                # Affichage du tableau avec filtres
                st.dataframe(display_data)

                # Créer un excel avec les colonnes inversées et le format pourcentage pour la variation
                excel_data = create_excel(data, selected_asset)

                st.download_button(
                    label="Télécharger les données (XLSX)",
                    data=excel_data,
                    file_name=f'{selected_asset}_{start_date_input}_{end_date_input}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                st.error(f"Aucune donnée n'a été récupérée pour {selected_asset}.")

    except Exception as e:
        st.error(f"Une erreur s'est produite lors de la récupération des données : {e}")
        import traceback
        st.error(f"Traceback détaillé: {traceback.format_exc()}")


# Affichage des données selon l'onglet sélectionné
with tab1:
    display_standard_asset_data(crypto_assets, "crypto")

with tab2:
    display_stock_data()  # Fonction spéciale pour les actions avec filtrage par secteur

with tab3:
    display_standard_asset_data(currency_assets, "currency")

with tab4:
    display_standard_asset_data(resource_assets, "resource")

with tab5:
    display_standard_asset_data(index_assets, "index")
